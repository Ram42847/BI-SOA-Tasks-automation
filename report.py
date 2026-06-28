#!/usr/bin/env python3
"""
SOA Weekly Task Status Report Generator
Project: MERP Services (id=75), Sunday-Saturday weeks, last 1 year tickets.

Usage:
  python report.py                         # Print HTML to stdout
  python report.py --output report.html    # Save HTML to file
  python report.py --send                  # Send via email (needs SMTP env vars)
  python report.py --week 2026-06-14       # Override week start date (must be Sunday)
"""

import argparse
import json
import os
import re
import smtplib
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import requests
from dotenv import load_dotenv

# ─── Config ──────────────────────────────────────────────────────────────────
load_dotenv()

OP_URL = os.getenv("OP_URL", "").rstrip("/")
OP_TOKEN = os.getenv("OP_TOKEN", "")
API_BASE = f"{OP_URL}/api/v3"
AUTH = ("apikey", OP_TOKEN)

PROJECT_ID = int(os.getenv("PROJECT_ID", "75"))
YEAR_LOOKBACK_DAYS = int(os.getenv("YEAR_LOOKBACK_DAYS", "365"))

# Type IDs used to split Merp tasks vs SOA Bugs in filters
API_BUG_TYPE_ID   = 19   # "API Bug"   → SOA Bugs section
API_STORY_TYPE_ID = 15   # "API Story" → Business tickets in bifurcation

# Environment custom field: exclude 'STAGE' bugs from SOA Bug counts
# customField9 = "Environment"; option id 22 = "STAGE"
ENV_STAGE_OPTION_ID = "22"

# All "open" status IDs (tasks still in progress)
OPEN_STATUS_IDS = {
    39, 44, 88, 36, 97, 42, 37, 40, 43, 46, 96, 59, 60, 61, 62, 63, 47, 98,
    93, 94, 95, 64, 89, 90, 91, 92, 82, 48, 49, 52, 50, 51, 53, 65, 54, 66,
    67, 68, 57, 69, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 1, 38, 99,
    100, 101,
}
BACKLOG_STATUS_IDS = {1}   # "To Do" = not yet aligned; WIP = open_end − backlog

CACHE_FILE = Path(__file__).parent / "weekly_data_cache.json"


# ─── OpenProject API helpers ─────────────────────────────────────────────────
def op_get(path, params=None):
    resp = requests.get(f"{API_BASE}{path}", auth=AUTH, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json()


def paginate(path, params=None, page_size=100, max_results=5000):
    results, offset = [], 1
    while True:
        p = dict(params or {}); p["pageSize"] = page_size; p["offset"] = offset
        data = op_get(path, p)
        elems = data.get("_embedded", {}).get("elements", [])
        results.extend(elems)
        if not elems or len(results) >= data.get("total", len(results)) or len(results) >= max_results:
            break
        offset += 1
    return results[:max_results]


def _link_id(links, key):
    href = (links.get(key) or {}).get("href") or ""
    m = re.search(r"/(\d+)$", href)
    return int(m.group(1)) if m else None


def _link_title(links, key, default=""):
    return (links.get(key) or {}).get("title") or default


def fetch_wps(extra_filters, page_size=100):
    base = [{"project": {"operator": "=", "values": [str(PROJECT_ID)]}}]
    params = {"filters": json.dumps(base + extra_filters)}
    return paginate("/work_packages", params, page_size=page_size)


# ─── Date helpers ─────────────────────────────────────────────────────────────
def last_completed_week(ref=None):
    """Return (sunday, saturday) of the most recently completed week."""
    today = ref or date.today()
    days_since_sat = (today.weekday() - 5) % 7   # Mon=0…Sat=5
    if days_since_sat == 0:
        days_since_sat = 7          # if today is Saturday, back-track a full week
    sat = today - timedelta(days=days_since_sat)
    sun = sat - timedelta(days=6)
    return sun, sat


def week_label(start, end):
    """e.g. '14 June - 20 June 2026'"""
    fmt = lambda d: d.strftime("%-d %B")
    return f"{fmt(start)} - {fmt(end)} {end.year}"


def short_label(start, end):
    """e.g. '14 June - 20 June'"""
    fmt = lambda d: d.strftime("%-d %b")
    return f"{fmt(start)} - {fmt(end)}"


def four_weeks(ref=None):
    """Return list of (start, end) tuples for 4 rolling weeks, oldest first."""
    latest_start, latest_end = last_completed_week(ref)
    return [(latest_start - timedelta(weeks=i), latest_end - timedelta(weeks=i))
            for i in range(3, -1, -1)]


def cache_key(start, end):
    return f"{start.isoformat()}_{end.isoformat()}"


# ─── Data computation for one week ───────────────────────────────────────────
def compute_week_data(week_start, week_end):
    """Fetch raw metrics for a single week (C is derived later in build_report_data)."""
    ws, we   = week_start.isoformat(), week_end.isoformat()
    cutoff   = (week_end - timedelta(days=YEAR_LOOKBACK_DAYS)).isoformat()

    open_ids_str = [str(i) for i in OPEN_STATUS_IDS]
    api_bug_str  = [str(API_BUG_TYPE_ID)]

    # Shared filter fragments
    open_filter    = {"status":      {"operator": "=",  "values": open_ids_str}}
    not_bug_filter = {"type":        {"operator": "!",  "values": api_bug_str}}
    bug_filter     = {"type":        {"operator": "=",  "values": api_bug_str}}
    not_stage_filter = {"customField9": {"operator": "!", "values": [ENV_STAGE_OPTION_ID]}}
    created_filter = {"createdAt":   {"operator": "<>d", "values": [ws, we]}}
    yr_filter      = {"createdAt":   {"operator": "<>d", "values": [cutoff, we]}}
    updated_filter = {"updatedAt":   {"operator": "<>d", "values": [ws, we]}}

    # ── Merp SOA Tasks ────────────────────────────────────────────────────────
    # B: opened this week, currently open, type != API Bug
    merp_opened_wps = fetch_wps([created_filter, open_filter, not_bug_filter])

    # Open end: currently open, type != API Bug, created in last 1 year
    merp_open_wps   = fetch_wps([yr_filter, open_filter, not_bug_filter])
    merp_open_end   = len(merp_open_wps)
    merp_backlog    = sum(1 for w in merp_open_wps
                          if _link_id(w.get("_links", {}), "status") in BACKLOG_STATUS_IDS)
    merp_wip        = merp_open_end - merp_backlog

    # ── SOA Bugs (type = API Bug, Environment != STAGE) ─────────────────────
    # B: opened this week, currently open, type = API Bug, env != stage
    bugs_opened_wps = fetch_wps([created_filter, open_filter, bug_filter, not_stage_filter])

    # Open end: currently open, type = API Bug, env != stage, created in last 1 year
    bugs_open_wps   = fetch_wps([yr_filter, open_filter, bug_filter, not_stage_filter])
    bugs_open_end   = len(bugs_open_wps)

    # ── Bifurcation ───────────────────────────────────────────────────────────
    # Business = API Story tickets in merp_opened; Internal = rest
    business = sum(1 for w in merp_opened_wps
                   if _link_id(w.get("_links", {}), "type") == API_STORY_TYPE_ID)
    internal = len(merp_opened_wps) - business

    # ── Developer tasks (open tickets, last 1 year, updated in week) ────────
    dev_wps   = fetch_wps([open_filter, yr_filter, updated_filter])
    dev_tasks = defaultdict(list)
    seen      = set()
    for wp in dev_wps:
        links       = wp.get("_links", {})
        wp_id       = wp.get("id")
        subject     = wp.get("subject", "")
        type_name   = _link_title(links, "type", "Task")
        assignee    = _link_title(links, "assignee", "")
        responsible = _link_title(links, "responsible", "")
        entry = {"id": wp_id, "subject": subject, "type": type_name}
        for person in {assignee, responsible}:
            if person and (person, wp_id) not in seen:
                dev_tasks[person].append(entry)
                seen.add((person, wp_id))

    # NOTE: merp_C and bugs_C are NOT returned here.
    # They are derived in build_report_data as: C = A + B - Open_End
    return {
        "merp_B":       len(merp_opened_wps),
        "merp_open":    merp_open_end,
        "merp_wip":     merp_wip,
        "merp_backlog": merp_backlog,
        "bugs_B":       len(bugs_opened_wps),
        "bugs_open":    bugs_open_end,
        "biz_opened":   business,
        "int_opened":   internal,
        "dev_tasks":    {k: sorted(v, key=lambda x: x["id"])
                         for k, v in sorted(dev_tasks.items())},
    }


# ─── Cache management ─────────────────────────────────────────────────────────
def load_cache():
    if CACHE_FILE.exists():
        return json.loads(CACHE_FILE.read_text())
    return {}


def save_cache(cache):
    CACHE_FILE.write_text(json.dumps(cache, indent=2))


# ─── Metric derivation for 4-week table ──────────────────────────────────────
def build_report_data(weeks, cache):
    """Return list of week dicts with all computed metrics.

    Historical weeks are loaded from cache as-is (values match the reference PDF).
    Current (fresh) week: A = previous week's Open_End from cache; C derived by formula.
    """
    rows      = []
    fresh_keys = set()

    for start, end in weeks:
        key = cache_key(start, end)
        if key in cache:
            d = cache[key]
        else:
            print(f"  Fetching data for {short_label(start, end)}...", file=sys.stderr)
            d = compute_week_data(start, end)
            cache[key] = d
            fresh_keys.add(key)
        rows.append({"start": start, "end": end, **d})

    for i, row in enumerate(rows):
        # Open_End displayed = raw cached value (already correct after cache seeding)
        row["merp_open_corrected"] = row["merp_open"]
        row["bugs_open_corrected"] = row["bugs_open"]

        # A = previous week's Open_End (from cache).
        # For the oldest week: back-compute from its own Open_End, B, and C.
        if i == 0:
            row["merp_A"] = row["merp_open"] - row["merp_B"] + row.get("merp_C", 0)
            row["bugs_A"] = row["bugs_open"] - row["bugs_B"] + row.get("bugs_C", 0)
        else:
            row["merp_A"] = rows[i - 1]["merp_open"]
            row["bugs_A"] = rows[i - 1]["bugs_open"]

        # For freshly fetched weeks: C = A + B − Open_End (no separate API call)
        key = cache_key(row["start"], row["end"])
        if key in fresh_keys:
            row["merp_C"] = row["merp_A"] + row["merp_B"] - row["merp_open"]
            row["bugs_C"] = row["bugs_A"] + row["bugs_B"] - row["bugs_open"]
            cache[key]["merp_C"] = row["merp_C"]
            cache[key]["bugs_C"] = row["bugs_C"]

    return rows


# ─── Excel detail sheets ──────────────────────────────────────────────────────
_EXCEL_HEADERS = (
    "ID", "Project", "Subject", "Type", "Status",
    "Author", "Assignee", "Accountable", "Updated on", "Created on",
)


def _fmt_dt(iso_str):
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return iso_str[:16]


def _wp_to_row(wp):
    links = wp.get("_links", {})
    return (
        wp.get("id", ""),
        _link_title(links, "project", ""),
        wp.get("subject", ""),
        _link_title(links, "type", ""),
        _link_title(links, "status", ""),
        _link_title(links, "author", ""),
        _link_title(links, "assignee", ""),
        _link_title(links, "responsible", ""),
        _fmt_dt(wp.get("updatedAt", "")),
        _fmt_dt(wp.get("createdAt", "")),
    )


def generate_excel_report(week_start, week_end, filepath):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        return

    ws_str  = week_start.isoformat()
    we_str  = week_end.isoformat()
    cutoff  = (week_end - timedelta(days=YEAR_LOOKBACK_DAYS)).isoformat()

    open_ids_str = [str(i) for i in OPEN_STATUS_IDS]
    api_bug_str  = [str(API_BUG_TYPE_ID)]

    open_filter      = {"status":      {"operator": "=",  "values": open_ids_str}}
    not_bug_filter   = {"type":        {"operator": "!",  "values": api_bug_str}}
    bug_filter       = {"type":        {"operator": "=",  "values": api_bug_str}}
    not_stage_filter = {"customField9": {"operator": "!", "values": [ENV_STAGE_OPTION_ID]}}
    yr_filter        = {"createdAt":   {"operator": "<>d", "values": [cutoff, we_str]}}
    updated_filter   = {"updatedAt":   {"operator": "<>d", "values": [ws_str, we_str]}}

    print("  Fetching tasks sheet (open, non-API Bug, updated in week)...", file=sys.stderr)
    tasks_wps = fetch_wps([open_filter, yr_filter, not_bug_filter, updated_filter])

    print("  Fetching bugs sheet (open API Bug, env != stage, updated in week)...", file=sys.stderr)
    bugs_wps  = fetch_wps([open_filter, yr_filter, bug_filter, not_stage_filter, updated_filter])

    hdr_font  = Font(bold=True)
    hdr_fill  = PatternFill("solid", fgColor="D9E1F2")
    col_widths = (8, 16, 60, 16, 20, 20, 20, 20, 18, 18)

    def write_sheet(ws, rows):
        ws.append(list(_EXCEL_HEADERS))
        for cell, w in zip(ws[1], col_widths):
            cell.font = hdr_font
            cell.fill = hdr_fill
            ws.column_dimensions[cell.column_letter].width = w
        for wp in rows:
            ws.append(list(_wp_to_row(wp)))

    start_tag = week_start.strftime("%-d%B")   # e.g. "21June"
    end_tag   = week_end.strftime("%-d%B")     # e.g. "27June"

    wb = Workbook()
    sh1 = wb.active
    sh1.title = f"{start_tag}-{end_tag}TasksDetails"
    write_sheet(sh1, tasks_wps)

    sh2 = wb.create_sheet(f"{start_tag}-{end_tag}BugsDetails")
    write_sheet(sh2, bugs_wps)

    wb.save(filepath)
    print(f"Excel report saved to {filepath}  "
          f"({len(tasks_wps)} tasks, {len(bugs_wps)} bugs)", file=sys.stderr)


# ─── HTML styling constants ───────────────────────────────────────────────────
HDR_BG     = "#2E6B73"   # dark teal  – column header rows
SECTION_BG = "#5F9EA0"   # cadet blue – section label rows

# +/- cell colours: light background + same-hue dark bold text (Excel palette)
PCT_AMBER_BG = "#FFEB9C";  PCT_AMBER_FG = "#9C6500"   # neutral  ≤5%
PCT_GREEN_BG = "#C6EFCE";  PCT_GREEN_FG = "#375623"   # good direction
PCT_PINK_BG  = "#FFC7CE";  PCT_PINK_FG  = "#9C0006"   # bad direction

TH_STYLE = f"background:{HDR_BG};color:white;padding:6px 10px;font-weight:bold"


# ─── HTML generation ──────────────────────────────────────────────────────────
def pct_cell(curr, prev, good_up=True):
    """Colored-background +/- cell.  good_up=True → increase is the good direction (green)."""
    if prev == 0:
        return '<td style="text-align:center;padding:5px 8px">—</td>'
    delta = curr - prev
    pct   = abs(delta) / prev * 100
    arrow = "▲" if delta >= 0 else "▼"
    if pct <= 5:
        bg, fg = PCT_AMBER_BG, PCT_AMBER_FG
    elif (delta > 0) == good_up:
        bg, fg = PCT_GREEN_BG, PCT_GREEN_FG
    else:
        bg, fg = PCT_PINK_BG, PCT_PINK_FG
    return (f'<td style="background:{bg};color:{fg};font-weight:bold;white-space:nowrap;'
            f'text-align:center;padding:5px 8px">'
            f'{arrow} {pct:.1f}%</td>')


def num(n):
    return f'<td style="text-align:right;padding:5px 10px">{n}</td>'


def build_html(rows):
    labels = [short_label(r["start"], r["end"]) for r in rows]
    curr   = rows[-1]
    prev   = rows[-2]

    # ── Key Highlights ────────────────────────────────────────────────────────
    closed_delta = curr["merp_C"] - prev["merp_C"]
    closed_pct   = abs(closed_delta) / prev["merp_C"] * 100 if prev["merp_C"] else 0
    closed_arrow = "▲" if closed_delta >= 0 else "▼"

    bugs_delta = curr["bugs_open_corrected"] - prev["bugs_open_corrected"]
    bugs_pct   = abs(bugs_delta) / prev["bugs_open_corrected"] * 100 if prev["bugs_open_corrected"] else 0
    bugs_arrow = "▲" if bugs_delta >= 0 else "▼"

    def hl_color(delta, pct, good_is_positive=True):
        if pct <= 5:
            return "#f39c12"
        good = (delta > 0) == good_is_positive
        return "#27ae60" if good else "#e74c3c"

    closed_color = hl_color(closed_delta, closed_pct, good_is_positive=True)
    bugs_color   = hl_color(bugs_delta,   bugs_pct,   good_is_positive=False)

    closed_word = "increased" if closed_delta >= 0 else "decreased"
    bugs_word   = "increased" if bugs_delta >= 0   else "decreased"

    highlight_html = f"""
    <ul>
      <li>The total number of <strong>closed tasks</strong> during this week has <strong>{closed_word}</strong> by
          <span style="color:{closed_color};font-weight:bold">{closed_arrow} {closed_pct:.1f}%</span>
          compared to the previous week's count.</li>
      <li>The total count of <strong>open bugs</strong> has <strong>{bugs_word}</strong> by
          <span style="color:{bugs_color};font-weight:bold">{bugs_arrow} {bugs_pct:.1f}%</span>
          compared to the previous week's count.</li>
    </ul>"""

    # ── Table header (shared between summary and bifurcation tables) ──────────
    th = "".join(f'<th style="{TH_STYLE};text-align:center">{l}</th>' for l in labels)
    header = f"""
    <tr>
      <th style="{TH_STYLE};text-align:left">Report Summary</th>
      {th}
      <th style="{TH_STYLE};text-align:center">(+/- Last Week)</th>
    </tr>"""

    def section_row(label):
        return (f'<tr><td colspan="6" style="background:{SECTION_BG};color:white;'
                f'font-weight:bold;padding:5px 10px">{label}</td></tr>')

    def data_row(label, vals, prev_val, curr_val, indent=False, good_up=True):
        cells = "".join(num(v) for v in vals)
        pad   = "padding:5px 10px;" + ("padding-left:20px" if indent else "")
        return (f'<tr><td style="{pad}">{label}</td>'
                f'{cells}{pct_cell(curr_val, prev_val, good_up)}</tr>')

    # ── MERP SOA Tasks section ────────────────────────────────────────────────
    # good_up direction: Opened↓=good, Closed↑=good, OpenEnd↓=good, WIP↑=good, Backlog↓=good
    merp_rows_html = (
        section_row("Merp SOA Tasks")
        + data_row("Open: Beginning of this week (A)",
                   [r["merp_A"] for r in rows],
                   prev["merp_A"], curr["merp_A"], good_up=False)
        + data_row("Opened During this week (B)",
                   [r["merp_B"] for r in rows],
                   prev["merp_B"], curr["merp_B"], good_up=False)
        + data_row("Closed During this week (C)",
                   [r["merp_C"] for r in rows],
                   prev["merp_C"], curr["merp_C"], good_up=True)
        + data_row("Open: End of this week (A + B - C)",
                   [r["merp_open_corrected"] for r in rows],
                   prev["merp_open_corrected"], curr["merp_open_corrected"], good_up=False)
        + '<tr><td colspan="6" style="padding:3px"></td></tr>'
        + data_row("Items in Under Development (WIP)",
                   [r["merp_wip"] for r in rows],
                   prev["merp_wip"], curr["merp_wip"], good_up=True)
        + data_row("Backlog (Yet to align)",
                   [r["merp_backlog"] for r in rows],
                   prev["merp_backlog"], curr["merp_backlog"], good_up=False)
    )

    # ── SOA Bugs section ──────────────────────────────────────────────────────
    # good_up direction: Opened↓=good, Closed↑=good, OpenEnd↓=good
    bugs_rows_html = (
        section_row("SOA Bugs - Open Project Tickets")
        + data_row("Open : Beginning of this week (A)",
                   [r["bugs_A"] for r in rows],
                   prev["bugs_A"], curr["bugs_A"], good_up=False)
        + data_row("Opened During this week (B)",
                   [r["bugs_B"] for r in rows],
                   prev["bugs_B"], curr["bugs_B"], good_up=False)
        + data_row("Closed During this week (C)",
                   [r["bugs_C"] for r in rows],
                   prev["bugs_C"], curr["bugs_C"], good_up=True)
        + data_row("Open: End of this week (A + B - C)",
                   [r["bugs_open_corrected"] for r in rows],
                   prev["bugs_open_corrected"], curr["bugs_open_corrected"], good_up=False)
    )

    summary_table = f"""
    <table border="1" cellspacing="0" cellpadding="0"
           style="border-collapse:collapse;font-size:13px;min-width:700px">
      {header}
      {merp_rows_html}
      <tr><td colspan="6" style="padding:6px"></td></tr>
      {bugs_rows_html}
    </table>"""

    # ── Bifurcation table ─────────────────────────────────────────────────────
    total_vals = [r["merp_B"] for r in rows]
    biz_vals   = [r["biz_opened"] for r in rows]
    int_vals   = [r["int_opened"] for r in rows]

    bif_table = f"""
    <table border="1" cellspacing="0" cellpadding="0"
           style="border-collapse:collapse;font-size:13px;min-width:700px;margin-top:20px">
      {header}
      {section_row("Merp SOA Tasks")}
      {data_row("Total Opened Tickets during this week (A=B+C)",
                total_vals, prev["merp_B"], curr["merp_B"], good_up=True)}
      <tr><td colspan="6" style="padding:3px"></td></tr>
      {data_row("Business Tickets (B)",
                biz_vals, prev["biz_opened"], curr["biz_opened"], good_up=True)}
      {data_row("Internally Opened Tickets (C)",
                int_vals, prev["int_opened"], curr["int_opened"], good_up=True)}
    </table>"""

    # ── Developer tasks table ─────────────────────────────────────────────────
    curr_week_label = week_label(curr["start"], curr["end"])
    dev_rows_html = ""
    dev_data = curr.get("dev_tasks", {})
    for developer, tasks in sorted(dev_data.items()):
        if not tasks:
            continue
        first = True
        for t in tasks:
            prefix = f'{t["type"]} #{t["id"]}'
            dev_cell = (f'<td rowspan="{len(tasks)}" style="vertical-align:top;'
                        f'padding:5px 10px;font-weight:bold">{developer}</td>'
                        if first else "")
            dev_rows_html += (
                f'<tr>{dev_cell}'
                f'<td style="padding:4px 10px">{prefix}: {t["subject"]}</td></tr>'
            )
            first = False

    dev_table = f"""
    <table border="1" cellspacing="0" cellpadding="0"
           style="border-collapse:collapse;font-size:13px;min-width:700px;margin-top:20px">
      <tr>
        <th style="{TH_STYLE};text-align:left;width:160px">Developer</th>
        <th style="{TH_STYLE};text-align:center">{curr_week_label}</th>
      </tr>
      <tr>
        <td style="padding:5px 10px"></td>
        <th style="background:{SECTION_BG};color:white;padding:5px 10px;text-align:center">Tasks</th>
      </tr>
      {dev_rows_html}
    </table>"""

    # ── Full email body ───────────────────────────────────────────────────────
    period = f"{short_label(curr['start'], curr['end'])} {curr['end'].year}"
    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"></head>
<body style="font-family:Arial,sans-serif;font-size:13px;color:#222;max-width:900px">

<p>Hi All,</p>
<p>Please find the Weekly Task Report of BI-SOA from
   <strong>{period}</strong>.</p>

<p><strong>Key Highlights :</strong></p>
{highlight_html}

{summary_table}

<p style="margin-top:20px">Please find below the bifurcation of total tickets opened :-</p>
{bif_table}

{dev_table}

<br>
<p>Please refer below for a detailed tracker.<br>
<a href="{OP_URL}/projects/{PROJECT_ID}/work_packages">BI-SOA Detailed Tracker</a></p>

<br>
<p><em>Thanks &amp; Regards,<br>
BI SOA Automation<br>
IndiaMart Intermesh Limited</em></p>

</body>
</html>"""


# ─── Email sending ────────────────────────────────────────────────────────────
def send_email(html, subject):
    smtp_host = os.getenv("SMTP_HOST", "")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER", "")
    smtp_pass = os.getenv("SMTP_PASS", "")
    mail_from = os.getenv("MAIL_FROM", smtp_user)
    mail_to   = os.getenv("MAIL_TO", "")
    mail_cc   = os.getenv("MAIL_CC", "")

    if not all([smtp_host, smtp_user, smtp_pass, mail_to]):
        print("ERROR: Set SMTP_HOST, SMTP_USER, SMTP_PASS, MAIL_TO in .env to send email.",
              file=sys.stderr)
        sys.exit(1)

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = mail_from
    msg["To"]      = mail_to
    if mail_cc:
        msg["Cc"]  = mail_cc

    msg.attach(MIMEText(html, "html"))

    recipients = [mail_to] + ([mail_cc] if mail_cc else [])
    with smtplib.SMTP(smtp_host, smtp_port) as s:
        s.starttls()
        s.login(smtp_user, smtp_pass)
        s.sendmail(mail_from, recipients, msg.as_string())
    print(f"Email sent to {mail_to}", file=sys.stderr)


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="SOA Weekly Report Generator")
    parser.add_argument("--output", help="Save HTML to this file")
    parser.add_argument("--send", action="store_true", help="Send email")
    parser.add_argument("--week", help="Override week start date YYYY-MM-DD (must be Sunday)")
    parser.add_argument("--no-cache", action="store_true", help="Re-fetch all weeks (ignore cache)")
    parser.add_argument("--excel", metavar="FILE", help="Save detail sheets to this .xlsx file")
    args = parser.parse_args()

    if not OP_URL or not OP_TOKEN:
        print("ERROR: OP_URL and OP_TOKEN must be set in .env", file=sys.stderr)
        sys.exit(1)

    if args.week:
        week_start_override = datetime.strptime(args.week, "%Y-%m-%d").date()
        week_end_override   = week_start_override + timedelta(days=6)
        weeks = [(week_start_override - timedelta(weeks=i),
                  week_end_override   - timedelta(weeks=i))
                 for i in range(3, -1, -1)]
    else:
        weeks = four_weeks()
    print(f"Generating report for weeks:", file=sys.stderr)
    for s, e in weeks:
        print(f"  {short_label(s, e)}", file=sys.stderr)

    cache = {} if args.no_cache else load_cache()

    # Always recompute current week (most recent)
    curr_key = cache_key(*weeks[-1])
    if curr_key in cache and not args.no_cache:
        print(f"Recomputing current week {short_label(*weeks[-1])}...", file=sys.stderr)
    cache.pop(curr_key, None)

    rows = build_report_data(weeks, cache)
    save_cache(cache)

    curr = rows[-1]
    subject = (f"BI Weekly Report :BI SOA Task Status Report : "
               f"{short_label(curr['start'], curr['end'])} {curr['end'].year}")

    html = build_html(rows)

    if args.output:
        Path(args.output).write_text(html)
        print(f"Report saved to {args.output}", file=sys.stderr)
    elif not args.send:
        print(html)

    if args.send:
        send_email(html, subject)

    if args.excel:
        generate_excel_report(curr["start"], curr["end"], args.excel)


if __name__ == "__main__":
    main()
